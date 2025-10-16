import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule

# Datos reales de las universidades
instagram_data = {
    'Institution': ['NYU', 'Columbia', 'Rutgers', 'Brandeis', 'Yeshiva', 'Maryland*'],
    'Followers': [593000, 457000, 124000, 25000, 15000, 4932],
    'Posts': [2613, 'N/A', 'N/A', 2965, 2260, 1258],
    'Market_Position': ['Leader', 'Premium', 'Challenger', 'Closest Peer', 'Current', 'Admissions Only']
}

# Datos reales de engagement por plataforma
platform_metrics = {
    'Platform': ['Instagram', 'TikTok', 'LinkedIn', 'Facebook', 'Twitter'],
    'Current_Rate': [1.5, 0, 1.2, 0.9, 0.8],
    'Benchmark': [2.99, 4.80, 2.95, 2.97, 2.61],
    'Gap': [-1.49, -4.80, -1.75, -2.07, -1.81],
    'Impact': ['High', 'Critical', 'Medium', 'Medium', 'Low']
}

# Datos reales de formatos de contenido
content_formats = {
    'Format': ['Instagram Reels', 'TikTok Videos', 'Static Posts', 'Carousel Posts', 'Live Content'],
    'Engagement_Rate': [1.99, 4.80, 0.80, 1.20, 3.50],
    'Completion_Rate': [85, 92, 'N/A', 65, 45],
    'Growth_Potential': ['Very High', 'Critical', 'Low', 'Medium', 'High']
}

# Datos reales de iniciativas estratégicas
strategic_initiatives = {
    'Initiative': ['TikTok Launch', 'Video Production', 'Team Expansion', 'Analytics System', 'Content Strategy'],
    'Required_Resources': [45000, 75000, 120000, 35000, 25000],
    'Timeline_Days': [90, 120, 60, 45, 30],
    'Expected_ROI': [285, 180, 150, 125, 200]
}

# Crear Excel con múltiples hojas
with pd.ExcelWriter('YU_Research_Documentation/03_Social_Media_Analysis/social_media_metrics.xlsx', engine='openpyxl') as writer:
    # Hoja 1: Instagram Performance
    df_instagram = pd.DataFrame(instagram_data)
    df_instagram.to_excel(writer, sheet_name='Instagram Performance', index=False)
    
    # Hoja 2: Platform Metrics
    df_platforms = pd.DataFrame(platform_metrics)
    df_platforms.to_excel(writer, sheet_name='Platform Metrics', index=False)
    
    # Hoja 3: Content Format Analysis
    df_content = pd.DataFrame(content_formats)
    df_content.to_excel(writer, sheet_name='Content Formats', index=False)
    
    # Hoja 4: Strategic Initiatives
    df_initiatives = pd.DataFrame(strategic_initiatives)
    df_initiatives.to_excel(writer, sheet_name='Strategic Initiatives', index=False)
    
    # Obtener el libro de trabajo
    workbook = writer.book
    
    # Dar formato a cada hoja
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Formato de encabezados
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
        # Alineación de celdas
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                
        # Reglas de formato condicional para columnas numéricas
        if sheet_name == 'Platform Metrics':
            worksheet.conditional_formatting.add('C2:C6',
                ColorScaleRule(start_type='min', start_color='FF9999',
                             end_type='max', end_color='99FF99'))
            
        if sheet_name == 'Content Formats':
            worksheet.conditional_formatting.add('B2:B6',
                ColorScaleRule(start_type='min', start_color='FF9999',
                             end_type='max', end_color='99FF99'))

# Crear hoja de resumen ejecutivo
summary_data = {
    'Metric': [
        'Total Instagram Followers',
        'Follower Gap vs Leader',
        'Current Engagement Rate',
        'Target Engagement Rate',
        'Required Investment',
        'Expected ROI (6 months)',
        'Growth Target (6 months)',
        'TikTok Growth Potential'
    ],
    'Value': [
        15000,
        -578000,
        '1.5%',
        '3.5%',
        '$300,000',
        '200%',
        '+67%',
        '2.28% weekly'
    ],
    'Status': [
        'Critical',
        'Critical',
        'Below Benchmark',
        'Achievable',
        'Required',
        'High',
        'Realistic',
        'High Priority'
    ]
}

df_summary = pd.DataFrame(summary_data)
with pd.ExcelWriter('YU_Research_Documentation/03_Social_Media_Analysis/executive_metrics.xlsx', engine='openpyxl') as writer:
    df_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
    
    workbook = writer.book
    worksheet = workbook['Executive Summary']
    
    # Formato de encabezados
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
    # Alineación de celdas
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
